import argparse
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
from openpyxl.chart import LineChart, Reference
import time

def init_driver():
    """Initialize the Selenium WebDriver."""
    return webdriver.Firefox()

def scrape_stock_data(driver, ticker):
    """
    Scrape stock data for the given ticker.

    Args:
        driver: The Selenium WebDriver.
        ticker: The stock ticker symbol.

    Returns:
        A pandas DataFrame with the scraped data.
    """
    base_url = f'https://stooq.pl/q/d/?s={ticker}&i=d'
    data = []
    page_number = 1
    while True:
        current_url = f"{base_url}&l={page_number}" if page_number > 1 else base_url
        driver.get(current_url)
        consent_if_needed(driver, page_number)
        
        # Wait for the dynamic content to load
        time.sleep(5)  # Adjust this delay as necessary
        
        rows = driver.find_elements(By.XPATH, "//table[@id='fth1']/tbody/tr")
        if not rows or len(data) >= 360:
            break
        for row in rows:
            cells = row.find_elements(By.TAG_NAME, 'td')
            if len(cells) == 9:  # Ensuring row has all data
                row_data = [cell.text for cell in cells[1:]]  # Skip first cell (index)
                data.append(row_data)
        page_number += 1
    
    columns = ['Date', 'Opening', 'High', 'Low', 'Closing Price', 'Change %', 'Change Nominal', 'Volume']
    return pd.DataFrame(data, columns=columns)

def consent_if_needed(driver, page_number):
    """
    Click the consent button on the page if it is found.

    Args:
        driver: The Selenium WebDriver.
        page_number: The current page number being scraped.
    """
    if page_number == 1:
        try:
            consent_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".fc-button.fc-cta-consent.fc-primary-button")))
            consent_button.click()
        except TimeoutException:
            pass  # Consent button not found or not clickable

def process_data_frame(df):
    """
    Process the DataFrame by converting date formats and data types.

    Args:
        df: The pandas DataFrame to process.
    """
    df['Date'] = df['Date'].apply(translate_date)
    df['Closing Price'] = pd.to_numeric(df['Closing Price'].str.replace(',', '.'), errors='coerce')

def save_to_excel(df, ticker):
    """
    Save the DataFrame to an Excel file and add charts based on the data.

    Args:
        df: The pandas DataFrame to save.
        ticker: The stock ticker symbol for naming the file.
    
    Returns:
        The filename of the saved Excel file.
    """
    excel_filename = f'scraped_stock_data_{ticker}_360_days.xlsx'
    df.to_excel(excel_filename, sheet_name='Stock Data', index=False)
    
    wb = openpyxl.load_workbook(excel_filename)
    ws_data = wb['Stock Data']
    ws_charts = wb.create_sheet('Charts')

    for i, days in enumerate([30, 180, 360], start=1):
        start_row = max(2, ws_data.max_row - days + 1)
        end_row = ws_data.max_row
        data = Reference(ws_data, min_col=5, min_row=start_row, max_row=end_row)
        categories = Reference(ws_data, min_col=1, min_row=start_row, max_row=end_row)

        chart = LineChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = f'Closing Prices Last {days} Days'
        ws_charts.add_chart(chart, f'A{15 * (i - 1) + 1}')

    wb.save(excel_filename)
    return excel_filename

def translate_date(date_str):
    """
    Translate Polish date string to English.

    Args:
        date_str: The date string in Polish.

    Returns:
        The date string in "YYYY-MM-DD" format.
    """
    polish_to_english = {
        'sty': '01', 'lut': '02', 'mar': '03', 'kwi': '04',
        'maj': '05', 'cze': '06', 'lip': '07', 'sie': '08',
        'wrz': '09', 'paź': '10', 'lis': '11', 'gru': '12'
    }
    day, month_abbr, year = date_str.split()
    month = polish_to_english[month_abbr.lower()]
    return f"{year}-{month}-{day.zfill(2)}"

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Scrape stock data')
    parser.add_argument('ticker', type=str, help='Stock ticker symbol')
    args = parser.parse_args()

    driver = init_driver()
    data_df = scrape_stock_data(driver, args.ticker)
    driver.quit()

    process_data_frame(data_df)
    excel_filename = save_to_excel(data_df, args.ticker)
    print(f"Excel file with charts saved as {excel_filename}.")
